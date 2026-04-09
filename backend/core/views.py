from collections import defaultdict
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import make_password, check_password
from .models import Faculty, Subject, Student, CourseOutcome, COMark, AttainmentLevel, POPSO, COPSOMap
from rest_framework.decorators import api_view
from rest_framework.response import Response
import csv
import io
import openpyxl
from django.http import HttpResponse
import json
import pandas as pd
from .models import COConfiguration
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from .models import POPSO
from .models import COConfiguration, POPSO, Subject, COPSOMap
from .models import COConfiguration, POPSO, Subject
from .models import Faculty
from .models import COPSOMap, POPSO

@csrf_exempt
def register_faculty(request):
    if request.method == "POST":
        try:
            if not request.body:
                return JsonResponse({"error": "Empty request body"}, status=400)

            data = json.loads(request.body)

            full_name = data.get("full_name")
            email = data.get("email")
            phone = data.get("phone")
            designation = data.get("designation")
            # Capture the branch from the request
            branch = data.get("branch")
            password = data.get("password")

            # Check missing fields (branch is optional for non-HODs, so not in 'all')
            if not all([full_name, email, phone, designation, password]):
                return JsonResponse({"error": "All fields are required"}, status=400)

            # Check duplicate email
            if Faculty.objects.filter(email=email).exists():
                return JsonResponse({"error": "Email already registered"}, status=400)

            # Logic to ensure branch is only saved if designation is HOD
            # This matches your frontend's .toUpperCase() logic
            final_branch = branch 
            # Create user with the new branch field
            faculty = Faculty.objects.create(
                full_name=full_name,
                email=email,
                phone=phone,
                designation=designation,
                branch=final_branch,
                password=make_password(password)
            )

            # Return the branch in the response so frontend can save it immediately
            return JsonResponse({
                "message": "Registration successful",
                "id": faculty.id,
                "full_name": faculty.full_name,
                "designation": faculty.designation,
                "branch": faculty.branch,
                "email": faculty.email,
                "phone": faculty.phone
            }, status=201)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
def login_faculty(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            email = data.get("email")
            password = data.get("password")
            login_type = data.get("login_type")  # HOD or FACULTY

            if not email or not password or not login_type:
                return JsonResponse(
                    {"error": "Email, password and login type are required"},
                    status=400
                )

            try:
                faculty = Faculty.objects.get(email=email)
            except Faculty.DoesNotExist:
                return JsonResponse({"error": "User not registered"}, status=400)

            # Check password
            if not check_password(password, faculty.password):
                return JsonResponse({"error": "Invalid password"}, status=400)

                        # 🔥 ROLE CHECK LOGIC (FIXED)
            designation = faculty.designation.strip().upper()

            if login_type == "HOD":
                if designation != "HOD":
                    return JsonResponse(
                        {"error": "Access denied. Only HOD can login here."},
                        status=403
                    )

            elif login_type == "PRINCIPAL":
                if designation != "PRINCIPAL":
                    return JsonResponse(
                        {"error": "Access denied. Only Principal can login here."},
                        status=403
                    )

            elif login_type == "FACULTY":
                if designation not in ["PROFESSOR", "HOD"]:
                    return JsonResponse(
                        {"error": "Access denied."},
                        status=403
                    )

            return JsonResponse({
                "message": "Login successful",
                "faculty_id": faculty.id,
                "faculty_name": faculty.full_name,
                "faculty_designation": faculty.designation,
                "faculty_branch": faculty.branch,
                "email": faculty.email,
                "phone": faculty.phone
            }, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def add_subject(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            faculty_id = data.get("faculty_id")
            subject_code = data.get("subject_code")
            subject_name = data.get("subject_name")
            semester = data.get("semester")
            branch = data.get("branch")
            session = data.get("session")
            batch = data.get("batch")

            faculty = Faculty.objects.get(id=faculty_id)

            if faculty.designation.strip().upper() != "HOD":
                return JsonResponse(
                    {"error": "Only HOD can add subjects"},
                    status=403
                )

            subject = Subject.objects.create(
                subject_code=subject_code,
                subject_name=subject_name,
                semester=semester,
                branch=branch,
                session=session,
                batch=batch,
                created_by=faculty
            )

            return JsonResponse({"message": "Subject added"}, status=201)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def get_hod_subjects(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            faculty_id = data.get("faculty_id")  # ✅ FIXED (moved outside)

            faculty = Faculty.objects.get(id=faculty_id)
            subjects = Subject.objects.filter(created_by=faculty)

            subject_list = []

            for subject in subjects:

                # ✅ GET CO DATA
                configs = COConfiguration.objects.filter(
                    subject=subject,
                )

                co_list = [
                    {
                        "co_number": co.co_number,
                        "co_description": co.statement
                    }
                    for co in configs
                ]

                # ✅ GET MARKS
                marks = COMark.objects.filter(
                    subject=subject
                ).select_related("student")

                student_map = {}

                for entry in marks:
                    student_id = entry.student.id

                    if student_id not in student_map:
                        student_map[student_id] = {
                            "name": entry.student.full_name,
                            "reg_no": entry.student.registration_number
                        }

                    if entry.co_number:
                        student_map[student_id][f"co{entry.co_number}"] = entry.marks

                subject_list.append({
                    "id": subject.id,
                    "subject_code": subject.subject_code,
                    "subject_name": subject.subject_name,
                    "semester": subject.semester,
                    "branch": subject.branch,
                    "session": subject.session,
                    "batch": subject.batch,
                    "is_active": subject.is_active,

                    # 🔥 IMPORTANT
                    "co_data": co_list,
                    "has_co": len(co_list) > 0,
                    "students": list(student_map.values())
                })

            return JsonResponse({"subjects": subject_list}, status=200)

        except Exception as e:
            print("ERROR:", str(e))
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def update_subject(request, subject_id):
    if request.method == "PUT":
        try:
            data = json.loads(request.body)

            faculty_id = data.get("faculty_id")

            faculty = Faculty.objects.get(id=faculty_id)

            if faculty.designation.strip().upper() != "HOD":
                return JsonResponse(
                    {"error": "Only HOD can edit subjects"},
                    status=403
                )

            subject = Subject.objects.get(id=subject_id)

            subject.subject_code = data.get("subject_code", subject.subject_code)
            subject.subject_name = data.get("subject_name", subject.subject_name)
            subject.semester = data.get("semester", subject.semester)
            subject.session = data.get("session", subject.session)
            subject.batch = data.get("batch", subject.batch)
            subject.is_active = data.get("is_active", subject.is_active)

            subject.save()

            return JsonResponse({"message": "Subject updated"}, status=200)

        except Subject.DoesNotExist:
            return JsonResponse({"error": "Subject not found"}, status=404)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def upload_students_csv(request):
    if request.method == "POST":
        try:
            session = request.POST.get("session")
            batch = request.POST.get("batch")
            semester = int(request.POST.get("semester"))
            branch = request.POST.get("branch")

            csv_file = request.FILES.get("file")

            if not csv_file:
                return JsonResponse({"error": "No file uploaded"}, status=400)

            decoded_file = csv_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            students_created = 0

            for row in reader:
                full_name = row.get("full_name")
                registration_number = row.get("registration_number")
                email = row.get("email")

                if not full_name or not registration_number:
                    continue

                if Student.objects.filter(
                    registration_number=registration_number.strip(),
                    batch=batch,
                    session=session,
                    # branch=branch
                ).exists():
                    continue

                Student.objects.create(
                    full_name=full_name.strip(),
                   registration_number=registration_number.strip(),
                    email=email.strip() if email else None,
                    branch=branch,
                    batch=batch,
                    semester=semester,
                    session=session
                )

                students_created += 1

            return JsonResponse({
                "message": f"{students_created} students uploaded successfully"
            })

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST allowed"}, status=405)


@csrf_exempt
def upload_faculty_csv(request):
    if request.method == "POST":
        try:
            csv_file = request.FILES.get("file")

            if not csv_file:
                return JsonResponse({"error": "No file uploaded"}, status=400)

            decoded_file = csv_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            faculty_created = 0

            for row in reader:
                full_name = row.get("full_name")
                email = row.get("email")
                designation = row.get("designation")
                branch = row.get("branch")

                # ✅ Only required fields (same as UI)
                if not all([full_name, email, designation]):
                    continue

                # ✅ Skip duplicate
                if Faculty.objects.filter(email=email.strip()).exists():
                    continue

                Faculty.objects.create(
                    full_name=full_name.strip(),
                    email=email.strip(),
                    designation=designation.strip().upper(),
                    branch=branch.strip().upper() if branch else "",
                    phone="",  # optional blank
                    password=make_password("123456")  # default password
                )

                faculty_created += 1

            return JsonResponse({
                "message": f"{faculty_created} faculty uploaded successfully"
            })

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST allowed"}, status=405)
@csrf_exempt
def get_students(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            session = data.get("session")
            batch = data.get("batch")
            semester = data.get("semester")
            branch = data.get("branch")
            if not all([session, batch, semester, branch]):
                return JsonResponse({"students": []}, status=200)
            students = Student.objects.filter(
                session=session,
                batch=batch,
                semester=int(semester),
                branch__iexact=branch
            )

            student_list = []

            for student in students:
                student_list.append({
                    "id": student.id,
                    "full_name": student.full_name,
                    "registration_number": student.registration_number,
                    "email": student.email,         
                    "semester": student.semester,
                })

            return JsonResponse({"students": student_list}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Only POST allowed"}, status=405)

@csrf_exempt
def add_course_outcome(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            batch = data.get("batch")
            session = data.get("session")
            semester = data.get("semester")
            branch = data.get("branch")
            faculty_id = data.get("faculty_id")
            subject_id = data.get("subject_id")

            # ✅ Basic validation
            if not all([batch, session, semester, faculty_id, subject_id]):
                return JsonResponse(
                    {"error": "All fields are required"},
                    status=400
                )

            # ✅ Safe faculty fetch
            try:
                faculty = Faculty.objects.get(id=faculty_id)
            except Faculty.DoesNotExist:
                return JsonResponse({"error": "Invalid faculty"}, status=400)

            # ✅ Safe subject fetch
            try:
                subject = Subject.objects.get(id=subject_id)
            except Subject.DoesNotExist:
                return JsonResponse({"error": "Invalid subject"}, status=400)

            # ✅ Prevent duplicate CO
            existing_co = CourseOutcome.objects.filter(
                subject=subject,
                batch=batch,
                semester=int(semester),
                session=session
            ).first()

            if existing_co:
                return JsonResponse({
                    "error": "Course Outcome already exists for this subject"
                }, status=400)

            # ✅ Create CO
            co = CourseOutcome.objects.create(
                faculty=faculty,
                subject=subject,
                batch=batch,
                session=session,
                semester=int(semester),
                branch=branch
            )

            return JsonResponse({
                "message": "Course Outcome added successfully",
                "id": co.id
            }, status=201)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def get_subjects_for_co(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            branch = data.get("branch")
            semester = int(data.get("semester"))
            batch = data.get("batch")
            session = data.get("session")
            faculty_id = data.get("faculty_id")   # 🔥 ADD THIS

            subjects = Subject.objects.filter(
                branch__iexact=branch,
                semester=semester,
                batch__iexact=batch,
                session__iexact=session,
                is_active=True
            )

            subject_list = []

            for subject in subjects:
                # 🔥 CHECK CO PER FACULTY
                has_co = COConfiguration.objects.filter(
                    subject=subject,
                    faculty_id=faculty_id
                ).exists()

                subject_list.append({
                    "id": subject.id,
                    "subject_code": subject.subject_code,
                    "subject_name": subject.subject_name,
                    "has_co": has_co   # 🔥 ADD THIS
                })

            return JsonResponse({"subjects": subject_list}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def save_co_marks(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            marks = data.get("marks", [])
            subject_id = data.get("subject_id")
            branch = data.get("branch")
            batch = data.get("batch")
            semester = data.get("semester")
            session = data.get("session")

            if not marks:
                return JsonResponse({"error": "No marks provided"}, status=400)

            for item in marks:
                student_id = item.get("student_id")
                co_number = item.get("co_number")
                mark_value = item.get("marks")

                COMark.objects.update_or_create(
                 student_id=student_id,
                 subject_id=subject_id,
                 co_number=co_number,
                 defaults={
                      "marks": mark_value,
                      "branch": branch,
                      "batch": batch,
                      "semester": semester
    }
)

            return JsonResponse({"message": "Marks saved successfully"}, status=200)

        except Exception as e:
            print("SAVE ERROR:", str(e))
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def get_co_marks(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            branch = data.get("branch")
            subject_id = data.get("subject_id")

            co_marks = COMark.objects.select_related(
                "student", "subject"
            )

            if branch:
                co_marks = co_marks.filter(branch__iexact=branch)

            if subject_id:
                co_marks = co_marks.filter(subject_id=subject_id)

            result = []
            for m in co_marks:
                result.append({
                    "student_name": m.student.full_name,
                    "reg_no": m.student.registration_number,
                    "subject_code": m.subject.subject_code,
                    "co_number": m.co_number,
                    "marks": m.marks,
                })

            return JsonResponse(result, safe=False)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "POST only"}, status=405)

@csrf_exempt
def get_branch_semester(request):
    faculty_id = request.GET.get("faculty_id")

    if not faculty_id:
        return JsonResponse([], safe=False)

    configs = COConfiguration.objects.select_related("subject").filter(
        faculty_id=faculty_id
    )

    data_map = {}

    for co in configs:
        subject = co.subject

        key = subject.id  # 🔥 UNIQUE SUBJECT

        if key not in data_map:
            data_map[key] = {
                "batch": subject.batch,
                "session": subject.session,
                "branch": subject.branch,
                "semester": subject.semester,
                "subject": subject.subject_name,
                "subject_id": subject.id,
            }

    return JsonResponse(list(data_map.values()), safe=False)
@csrf_exempt
def download_excel(request, branch, semester):
    
    subject_name = request.GET.get("subject")

    marks = COMark.objects.filter(
        branch__iexact=branch,
        semester=semester,
        subject__subject_name=subject_name
    ).select_related("student", "subject")

    # 🔷 FETCH CO DETAILS
    co_details = COConfiguration.objects.filter(
        subject__subject_name=subject_name,
        subject__branch__iexact=branch,
        subject__semester=semester
    ).order_by("co_number")

    wb = openpyxl.Workbook()
    ws = wb.active
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
)
    # =========================
    # 🔥 ADD HEADER SECTION
    # =========================

    ws.append(["Subject", subject_name])
    ws.append(["Branch", branch])
    ws.append(["Semester", semester])

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
      for cell in row:
        cell.font = bold_font

    ws.append([])

    ws.append(["Course Outcomes"])
    ws["A5"].font = Font(bold=True, size=14)

    # ✅ Write CO rows
    start_row = 6

    for co in co_details:
        ws.append([f"CO{co.co_number}", co.statement])

    end_row = start_row + len(co_details) - 1
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=2):
      for cell in row:
         cell.border = thin_border
         cell.alignment = Alignment(wrap_text=True)

    ws.append([])
    ws.append([])

    # =========================
    # EXISTING LOGIC (UNCHANGED)
    # =========================

    students = defaultdict(dict)
    co_set = set()

    for m in marks:
        reg = m.student.registration_number
        co = f"CO{m.co_number}"

        students[reg]["name"] = m.student.full_name
        students[reg]["registration_number"] = reg
        students[reg]["subject"] = m.subject.subject_name
        students[reg][co] = m.marks

        co_set.add(co)

    co_list = sorted(co_set)

    headers = ["Student Name", "Registration Number", "Subject"] + co_list + ["Total"]
    ws.append(headers)

    co_totals = {co: 0 for co in co_list}
    student_count = 0

    for student in students.values():

        row = [
            student.get("name"),
            student.get("registration_number"),
            student.get("subject")
        ]

        total = 0

        for co in co_list:
            mark = student.get(co, 0)
            row.append(mark)

            total += mark
            co_totals[co] += mark

        row.append(total)
        ws.append(row)

        student_count += 1
        

    # =========================
    # ANALYTICS (UNCHANGED)
    # =========================

    co_avg = {}
    co_has_data = {}
    for co in co_list:
        #check if any non-zero marks exist
        has_real_data = any(student.get(co, 0) > 0 for student in students.values())

        co_has_data[co] = has_real_data
        if student_count > 0 and has_real_data:

            co_avg[co] = round(co_totals[co] / student_count, 2)
        else:
            co_avg[co] = 0

    above_avg = {co: 0 for co in co_list}

    
    for student in students.values():
        for co in co_list:
             # skip fake CO (all zero case)
            if not co_has_data[co]:
                continue
            if student.get(co, 0) >= co_avg[co]:
                above_avg[co] += 1

    # 🔥 GET SUBJECT SESSION
    subject_obj = Subject.objects.filter(subject_name=subject_name,branch__iexact=branch,
    semester=semester).first()
    session = subject_obj.session if subject_obj else None
    levels = AttainmentLevel.objects.filter(session=session).order_by('-id').first()

    # 🔥 FETCH LEVELS ONCE

    if levels:
        level1 = levels.level1
        level2 = levels.level2
        level3 = levels.level3
    else:
        level1, level2, level3 = 50, 60, 70

    co_attainment = {}

    
    for co in co_list:
         #  handle no real data
        if not co_has_data[co]:
            co_attainment[co] = 0
            continue
        percentage = (above_avg[co] / student_count) * 100 if student_count else 0
        if percentage >= level3:
            level = 3
        elif percentage >= level2:
            level = 2
        elif percentage >= level1:
            level = 1
        else:
            level = 0

        final_attainment = round(0.9 * level, 2)

        co_attainment[co] = final_attainment

    ws.append([])

    avg_row = ["", "", "Average Marks"]
    for co in co_list:
        avg_row.append(co_avg[co])
    avg_row.append("")
    ws.append(avg_row)

    count_row = ["", "", "Students ≥ Avg"]
    for co in co_list:
        count_row.append(above_avg[co])
    count_row.append("")
    ws.append(count_row)

    attain_row = ["", "", "CO Attainment"]
    for co in co_list:
        attain_row.append(co_attainment[co])
    attain_row.append("")
    ws.append(attain_row)

    avg_attainment = round(sum(co_attainment.values()) / len(co_attainment), 2) if co_attainment else 0

    ws.append([])
    ws.append(["", "", "Average CO Attainment", avg_attainment])
    for row in ws.iter_rows(min_row=ws.max_row-3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
      for cell in row:
          cell.font = bold_font
          cell.alignment = center_align

    # =========================
    # RESPONSE
    # =========================

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f"attachment; filename={subject_name}_{branch}_Sem{semester}.xlsx"

    wb.save(response)
    for col in ws.columns:
      max_length = 0
      col_letter = col[0].column_letter

    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    ws.column_dimensions[col_letter].width = max_length + 2

    return response
@csrf_exempt
def principal_dashboard_stats(request):
    data = {
        "students": Student.objects.count(),
        "faculty": Faculty.objects.count(),
        "subjects": Subject.objects.count(),
    }
    return JsonResponse(data)
def get_all_students(request):
    if request.method == "GET":
        students = Student.objects.all()

        student_list = []

        for student in students:
            student_list.append({
                "id": student.id,
                "full_name": student.full_name,
                "registration_number": student.registration_number,
                "email": student.email,
                "branch": student.branch,
                "semester": student.semester,
            })

        return JsonResponse(student_list, safe=False)

    return JsonResponse({"error": "Only GET allowed"}, status=405)
def get_all_faculty(request):
    if request.method == "GET":
        faculty = Faculty.objects.all()

        faculty_list = []

        for f in faculty:
            faculty_list.append({
                "id": f.id,
                "full_name": f.full_name,
                "email": f.email,
                "designation": f.designation,
                "branch": f.branch,
                "phone": f.phone
            })

        return JsonResponse(faculty_list, safe=False)

    return JsonResponse({"error": "Only GET allowed"}, status=405)
@csrf_exempt
def get_all_subjects(request):
    if request.method == "GET":
        subjects = Subject.objects.all()

        subject_list = []
        for s in subjects:
            subject_list.append({
                "id": s.id,
                "subject_code": s.subject_code,
                "subject_name": s.subject_name,
                "branch": s.branch,
                "semester": s.semester,
                "session": s.session,
                "batch": s.batch,
                 "faculty_name": s.created_by.full_name 
            })

        return JsonResponse(subject_list, safe=False)

    return JsonResponse({"error": "Only GET allowed"}, status=405)
def principal_co(request):
    branch = request.GET.get("branch")
    semester = request.GET.get("semester")
    subject_id = request.GET.get("subject_id")

    subjects = Subject.objects.filter(
        branch__iexact=branch,
        semester=semester
    )

    if subject_id:
        subjects = subjects.filter(id=subject_id)

    data = []

    for sub in subjects:
        cos = COConfiguration.objects.filter(subject=sub)  # ✅ FIXED MODEL

        if cos.exists():
            data.append({
                "subject_name": sub.subject_name,
                "co": [
                    {
                        "id": co.id,
                        "co_number": f"CO{co.co_number}",   # ✅ formatted
                        "description": co.statement         # ✅ FIXED FIELD
                    }
                    for co in cos
                ]
            })

    return JsonResponse(data, safe=False)
def export_students_excel(request):
    import openpyxl
    from django.http import HttpResponse
    from .models import Student

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append(["Name", "Registration No", "Branch", "Semester"])

    branch = request.GET.get("branch")
    semester = request.GET.get("semester")

    students = Student.objects.all()

    if branch:
        students = students.filter(branch__iexact=branch)

    if semester:
        students = students.filter(semester=semester)

    for s in students:
        ws.append([
            s.full_name,
            s.registration_number,
            s.branch,
            s.semester
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=students.xlsx"

    wb.save(response)
    return response

def export_faculty_excel(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faculty"

    ws.append(["Name", "Email", "Designation", "Branch"])

    branch = request.GET.get("branch")

    faculty = Faculty.objects.all()

    if branch:
        faculty = faculty.filter(branch__iexact=branch)

    for f in faculty:
        ws.append([
            f.full_name,
            f.email,
            f.designation,
            f.branch
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=faculty.xlsx"

    wb.save(response)
    return response
def download_co_pdf(request):
    branch = request.GET.get("branch")
    semester = request.GET.get("semester")
    subject_id = request.GET.get("subject_id")

    subjects = Subject.objects.filter(
        branch__iexact=branch,
        semester=semester
    )

    if subject_id:
        subjects = subjects.filter(id=subject_id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=CO_Details.pdf"

    doc = SimpleDocTemplate(response)
    elements = []
    styles = getSampleStyleSheet()

    # TITLE
    elements.append(Paragraph("Course Outcome (CO) Details", styles["Title"]))
    elements.append(Spacer(1, 10))

    for sub in subjects:
        elements.append(Paragraph(f"<b>Subject:</b> {sub.subject_name}", styles["Heading2"]))
        elements.append(Paragraph(f"<b>Branch:</b> {sub.branch}", styles["Normal"]))
        elements.append(Paragraph(f"<b>Semester:</b> {sub.semester}", styles["Normal"]))
        elements.append(Spacer(1, 10))

        cos = COConfiguration.objects.filter(subject=sub)

        data = [["CO Number", "Description"]]

        for co in cos:
            data.append([f"CO{co.co_number}", co.statement])

        table = Table(data, colWidths=[100, 400])

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    return response
def export_subjects_excel(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subjects"

    ws.append(["Subject Code", "Subject Name", "Branch", "Semester", "Session", "Batch"])

    branch = request.GET.get("branch")
    semester = request.GET.get("semester")

    subjects = Subject.objects.all()

    if branch:
        subjects = subjects.filter(branch__iexact=branch)

    if semester:
        subjects = subjects.filter(semester=semester)

    for s in subjects:
        ws.append([
            s.subject_code,
            s.subject_name,
            s.branch,
            s.semester,
            s.session,
            s.batch
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=subjects.xlsx"

    wb.save(response)
    return response
def export_co_marks_excel(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CO Analytics"

    ws.append([
        "Student Name",
        "Reg No",
        "Subject Code",
        "CO Number",
        "Marks"
    ])

    branch = request.GET.get("branch")
    subject_id = request.GET.get("subject_id")

    co_marks = COMark.objects.select_related("student", "subject")

    if branch:
        co_marks = co_marks.filter(branch__iexact=branch)

    if subject_id:
        co_marks = co_marks.filter(subject_id=subject_id)

    for m in co_marks:
        ws.append([
            m.student.full_name,
            m.student.registration_number,
            m.subject.subject_code,
            m.co_number,
            m.marks
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=co_analytics.xlsx"

    wb.save(response)
    return response
@csrf_exempt
def get_co_analytics(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            branch = data.get("branch")
            subject_id = data.get("subject_id")

            marks = COMark.objects.filter(
                branch__iexact=branch,
                subject_id=subject_id
            ).select_related("student", "subject")

            from collections import defaultdict

            students = defaultdict(dict)
            co_set = set()

            for m in marks:
                reg = m.student.registration_number

                students[reg]["name"] = m.student.full_name
                students[reg]["registration_number"] = reg
                students[reg]["subject"] = m.subject.subject_name

                co_key = f"CO{m.co_number}"
                students[reg][co_key] = m.marks

                co_set.add(co_key)

            co_list = sorted(co_set)

            result = []
            for student in students.values():
                row = {
                    "name": student["name"],
                    "registration_number": student["registration_number"],
                    "subject": student["subject"],
                }

                total = 0
                for co in co_list:
                    mark = student.get(co, 0)
                    row[co] = mark
                    total += mark

                row["total"] = total
                result.append(row)

            return JsonResponse({
                "data": result,
                "co_list": co_list
            })

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "POST only"}, status=405)
@csrf_exempt
def course_attainment(request):
    try:
        # ✅ GET SESSION FROM FRONTEND
        session_param = request.GET.get("session")

        marks = COMark.objects.select_related("student", "subject")

        # ✅ FILTER MARKS BY SESSION (IMPORTANT)
        if session_param:
            marks = marks.filter(session=session_param)

        from collections import defaultdict

        subject_data = defaultdict(lambda: defaultdict(list))

        # Group marks
        for m in marks:
            key = (m.subject.subject_name, m.branch, m.semester)
            subject_data[key][f"CO{m.co_number}"].append(m.marks)

        results = []

        for (subject, branch, semester), co_dict in subject_data.items():

            attainment_list = []
            level_list = []

            # 🔥 GET TOTAL CO COUNT FROM CONFIG
            co_configs = COConfiguration.objects.filter(
                subject__subject_name=subject,
                subject__branch__iexact=branch,
                subject__semester=semester
            )

            total_co = co_configs.count()

            attainment_list = []

            #  LOOP THROUGH ALL COs (NOT JUST EXISTING ONES)
            for i in range(1, total_co + 1):
                co_key = f"CO{i}"
                marks_list = co_dict.get(co_key, [])
#  HANDLE EMPTY OR ALL-ZERO MARKS
                if not marks_list or all(m == 0 for m in marks_list):
                    attainment_list.append(0)
                    level_list.append(0)
                    continue

                total_students = len(marks_list)
                avg = sum(marks_list) / total_students

                above_avg = sum(1 for m in marks_list if m >= avg)

                percentage = (above_avg / total_students) * 100 if total_students else 0

                # 🔥 LEVEL FETCH (SESSION FIX APPLIED)
                subject_obj = Subject.objects.filter(
                    subject_name=subject,
                    branch__iexact=branch,
                    semester=semester
                ).first()

               
               # ✅ ALWAYS TAKE LATEST ATTAINMENT SESSION (CORRECT SOURCE)
                latest_session = AttainmentLevel.objects.order_by('-id').values_list('session', flat=True).first()

                session_val = latest_session

                levels = AttainmentLevel.objects.filter(session=session_val).order_by('-id').first()

                if levels:
                    level1 = levels.level1
                    level2 = levels.level2
                    level3 = levels.level3
                else:
                    level1, level2, level3 = 50, 60, 70

                # 🔥 DISCRETE LEVEL LOGIC (UNCHANGED)

                if percentage >= level3:
                    level = 3
                elif percentage >= level2:
                    level = 2
                elif percentage >= level1:
                    level = 1
                else:
                    level = 0

                course_attainment = round(0.9 * level, 2)
                attainment_list.append(course_attainment)
                level_list.append(level)

            final_attainment = sum(attainment_list) / len(attainment_list)
            final_level = round(sum(level_list) / len(level_list)) if level_list else 0

            results.append({
                "branch": branch,
                "subject": subject,
                "semester": semester,
                "session": latest_session,  #  FIXED
                "attainment": round(final_attainment, 2),
                "level": final_level
            })
            #  IF NO DATA, STILL RETURN SUBJECT STRUCTURE
        if not results:
            subjects = Subject.objects.all()

            for s in subjects:
                results.append({
                    "branch": s.branch,
                    "subject": s.subject_name,
                    "semester": s.semester,
                    "session": latest_session,
                    "attainment": 0,
                    "level": 0
        })

        return JsonResponse(results, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
from django.http import HttpResponse
import pandas as pd
from io import BytesIO
import json

@csrf_exempt
def export_coa(request):
    try:
        body = json.loads(request.body)
        data = body.get("data", [])

        df = pd.DataFrame(data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="COA")

        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=COA_Report.xlsx'

        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)
@csrf_exempt
def save_co_details(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            subject_id = data.get("subject_id")
            co_statements = data.get("co_statements")
            faculty_id = data.get("faculty_id")

            subject = Subject.objects.get(id=subject_id)
            faculty = Faculty.objects.get(id=faculty_id)

            # ✅ Prevent duplicate per faculty
            if COConfiguration.objects.filter(
                subject=subject,
                faculty=faculty
            ).exists():
                return JsonResponse({
                    "success": False,
                    "message": "CO details already added"
                }, status=400)

            # ✅ Save with faculty
            for i, statement in enumerate(co_statements, start=1):
                COConfiguration.objects.create(
                    subject=subject,
                    faculty=faculty,   # 🔥 IMPORTANT
                    co_number=i,
                    statement=statement
                )

            return JsonResponse({
                "success": True,
                "message": "CO details saved successfully"
            }, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
@csrf_exempt
def get_subjects_with_co(request):
    if request.method == "GET":
        try:
            faculty_id = request.GET.get("faculty_id")

            subjects = Subject.objects.filter(
                coconfiguration__faculty_id=faculty_id
            ).distinct()

            data = []
            for sub in subjects:
                total_cos = COConfiguration.objects.filter(subject=sub).count()

                filled_cos = COMark.objects.filter(
                    subject=sub,
                    marks__gt=0   #  ONLY COUNT REAL FILLED MARKS
                ).values('co_number').distinct().count()
                # 🔥 GET ALL CO NUMBERS
                all_cos = list(
                    COConfiguration.objects.filter(subject=sub)
                    .values_list("co_number", flat=True)
)

# 🔥 GET FILLED CO NUMBERS (marks > 0)
                filled_cos_list = list(
                    COMark.objects.filter(subject=sub, marks__gt=0)
                    .values_list("co_number", flat=True)
                    .distinct()
)

# 🔥 FIND MISSING COs
                missing_cos = [co for co in all_cos if co not in filled_cos_list]

                all_filled = (total_cos > 0 and total_cos == filled_cos)
                has_marks = filled_cos > 0
                

                data.append({
                    "id": sub.id,
                    "subject_code": sub.subject_code,
                    "subject_name": sub.subject_name,
                    "branch": sub.branch,
                    "semester": sub.semester,
                    "batch": sub.batch,
                    "session": sub.session,
                    "has_marks": has_marks,
                    "all_marks_filled": all_filled,
                    "missing_cos": missing_cos
                })

            return JsonResponse({"subjects": data}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
@csrf_exempt
def get_co_details(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            subject_id = data.get("subject_id")

            faculty_id = data.get("faculty_id")

            cos = COConfiguration.objects.filter(
               subject_id=subject_id,
               faculty_id=faculty_id
        ).values("co_number", "statement")

            return JsonResponse({
                "cos": list(cos)
            }, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
from django.db.models import Count
@csrf_exempt
def get_all_co_details(request):
    faculty_id = request.GET.get("faculty_id")

    if not faculty_id:
        return JsonResponse([], safe=False)

    configs = COConfiguration.objects.select_related("subject").filter(
        faculty_id=faculty_id
    )

    data = {}

    for co in configs:
        subject = co.subject

        key = subject.id

        if key not in data:
            data[key] = {
    "subject_id": subject.id,   #  ADD THIS
    "subject_code": subject.subject_code,
    "subject_name": subject.subject_name,
    "branch": subject.branch,
    "semester": subject.semester,
    "session": subject.session,
    "co_count": 0,
    
}
        data[key]["co_count"] += 1

    return JsonResponse(list(data.values()), safe=False)
@csrf_exempt
def get_co_by_subject(request):
    subject_id = request.GET.get("subject_id")
    faculty_id = request.GET.get("faculty_id")

    if not subject_id:
        return JsonResponse([], safe=False)

    #  HANDLE BOTH CASES
    if faculty_id:
        cos = COConfiguration.objects.filter(
            subject_id=subject_id,
            faculty_id=faculty_id
        ).order_by("co_number")
    else:
        #  fallback (for safety)
        cos = COConfiguration.objects.filter(
            subject_id=subject_id
        ).order_by("co_number")

    data = [
        {
            "id": co.id,
            "co_number": co.co_number,
            "co_description": co.statement
        }
        for co in cos
    ]

    return JsonResponse(data, safe=False)
@csrf_exempt
def update_co(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            co_id = data.get("id")
            statement = data.get("statement")
            faculty_id = data.get("faculty_id")

            co = COConfiguration.objects.get(id=co_id)

            #  SECURITY: only owner can update
            if str(co.faculty_id) != str(faculty_id):
                return JsonResponse({"error": "Unauthorized"}, status=403)

            co.statement = statement
            co.save()

            return JsonResponse({"message": "Updated successfully"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
@csrf_exempt
def add_co_single(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            subject_code = data.get("subject_code")
            co_number = data.get("co_number")
            statement = data.get("statement")
            faculty_id = data.get("faculty_id")

            subject = Subject.objects.get(subject_code=subject_code)
            faculty = Faculty.objects.get(id=faculty_id)

            #  Prevent duplicate CO number per faculty
            if COConfiguration.objects.filter(
                subject=subject,
                faculty=faculty,
                co_number=co_number
            ).exists():
                return JsonResponse({"error": "CO already exists"}, status=400)

            co=COConfiguration.objects.create(
                subject=subject,
                faculty=faculty,   # IMPORTANT
                co_number=co_number,
                statement=statement
            )
            # 🔥 ADD DEFAULT MARKS FOR NEW CO

            existing_marks = COMark.objects.filter(subject=subject)

            students_data = existing_marks.values(
                "student", "branch", "semester", "session"
            ).distinct()

            new_marks = []

            for data in students_data:
                new_marks.append(
                    COMark(
                        student_id=data["student"],
                        subject=subject,
                        co_number=co.co_number,
                        marks=0,
                        branch=data["branch"],
                        semester=data["semester"],
                        session=data["session"],
        )
    )

            COMark.objects.bulk_create(new_marks)

            return JsonResponse({"message": "CO added successfully"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
@csrf_exempt
def delete_co(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            co_id = data.get("id")
            faculty_id = data.get("faculty_id")

            co = COConfiguration.objects.get(id=co_id)
            subject = co.subject
            co_number = co.co_number

            #  SECURITY: only owner can delete
            if str(co.faculty_id) != str(faculty_id):
                return JsonResponse({"error": "Unauthorized"}, status=403)

            # delete config
            co.delete()

            #  delete related marks
            COMark.objects.filter(
                subject=subject,
                co_number=co_number
).delete()
            #  REORDER CO NUMBERS
            remaining_cos = COConfiguration.objects.filter(subject=subject).order_by('co_number')

            for index, c in enumerate(remaining_cos, start=1):
                if c.co_number != index:
                    old_number = c.co_number
                    c.co_number = index
                    c.save()

                #  ALSO UPDATE MARKS
                    COMark.objects.filter(
                        subject=subject,
                        co_number=old_number
                    ).update(co_number=index)

            return JsonResponse({"message": "Deleted successfully"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

def download_co_details(request):
    subject_code = request.GET.get("subject_code")
    faculty_id = request.GET.get("faculty_id")

    if not subject_code or not faculty_id:
        return HttpResponse("Missing subject_code or faculty_id", status=400)

    cos = COConfiguration.objects.filter(
        subject__subject_code=subject_code,
        faculty_id=faculty_id
    )

    if not cos.exists():
        return HttpResponse("No data found", status=404)

    subject = cos.first().subject

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{subject.subject_name}_{subject.branch}_Sem{subject.semester}_CO_Details.csv"'

    writer = csv.writer(response)

    writer.writerow(["Subject:", subject.subject_name])
    writer.writerow(["Branch:", subject.branch])
    writer.writerow(["Semester:", subject.semester])
    writer.writerow([])

    writer.writerow(["CO Number", "Statement"])

    for co in cos:
        writer.writerow([co.co_number, co.statement])

    return response
def get_sessions(request):
    try:
        from .models import Subject, AttainmentLevel

        #  ALL sessions (from Subject)
        subject_sessions = Subject.objects.values_list("session", flat=True)

        #  sessions with attainment
        attainment_sessions = AttainmentLevel.objects.values_list("session", flat=True)

        #  MERGE + UNIQUE
        all_sessions = set(subject_sessions) | set(attainment_sessions)

        return JsonResponse(sorted(list(all_sessions)), safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

from django.http import JsonResponse    
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def save_attainment(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            session = data.get("session")
            level1 = data.get("level1")
            level2 = data.get("level2")
            level3 = data.get("level3")

            if not all([session, level1, level2, level3]):
                return JsonResponse({"error": "All fields required"}, status=400)

            # 🔥 THIS IS THE FIX (UPDATE INSTEAD OF CREATE)
            obj, created = AttainmentLevel.objects.update_or_create(
                session=session,
                defaults={
                    "level1": level1,
                    "level2": level2,
                    "level3": level3
                }
            )

            return JsonResponse({"message": "Saved successfully"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    

@csrf_exempt
def get_attainment(request):
    try:
        session = request.GET.get("session")

        if not session:
            return JsonResponse({}, safe=False)

        #  FIX: always take latest entry
        data = AttainmentLevel.objects.filter(session=session).order_by('-id').first()

        if not data:
            return JsonResponse({}, safe=False)

        return JsonResponse({
            "level1": data.level1,
            "level2": data.level2,
            "level3": data.level3
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



@csrf_exempt
def save_po_pso(request):
    if request.method in ["POST", "PUT"]:
        try:
            data = json.loads(request.body)

            print("DATA RECEIVED:", data)

            branch = "CSE"   # TEMP (later replace with user branch)
            session = data.get("session")
            pos = data.get("pos", [])
            psos = data.get("psos", [])

            if not session:
                return JsonResponse({"error": "Session required"}, status=400)

            #  delete old records
            POPSO.objects.filter(branch=branch, session=session).delete()

            # Save PO
            for i, po in enumerate(pos):
                if po.strip():
                    POPSO.objects.create(
                        type="PO",
                        code=f"PO{i+1}",
                        description=po,
                        branch=branch,
                        session=session
                    )

            # Save PSO
            for i, pso in enumerate(psos):
                if pso.strip():
                    POPSO.objects.create(
                        type="PSO",
                        code=f"PSO{i+1}",
                        description=pso,
                        branch=branch,
                        session=session
                    )

            return JsonResponse({"message": "Saved successfully"})

        except Exception as e:
            print("ERROR:", str(e))
            return JsonResponse({"error": str(e)}, status=500)




def get_mapping_data(request):
    subject_id = request.GET.get("subject_id")

    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"error": "Invalid subject"}, status=400)

    
    cos = COConfiguration.objects.filter(subject=subject).order_by("co_number")

   
    pos = POPSO.objects.filter(
        branch=subject.branch,
        session=subject.session,
        type="PO"
    ).order_by("code")

    
    psos = POPSO.objects.filter(
        branch=subject.branch,
        session=subject.session,
        type="PSO"
    ).order_by("code")

    return JsonResponse({
        "cos": [
            {"id": c.id, "text": c.statement, "co_number": c.co_number}
            for c in cos
        ],
        "pos": [
            {"code": p.code, "description": p.description}
            for p in pos
        ],
        "psos": [
            {"code": p.code, "description": p.description}
            for p in psos
        ]
    })

@csrf_exempt
def save_co_po_pso(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            subject_id = data.get("subject_id")
            mappings = data.get("mappings")

            subject = Subject.objects.get(id=subject_id)

            #  DELETE OLD (for update)
            COPSOMap.objects.filter(subject=subject).delete()

            for m in mappings:
                co_obj = COConfiguration.objects.get(id=m["co"])

                COPSOMap.objects.create(
                    subject=subject,
                    co=co_obj,
                    po_mapping=m["po"],
                    pso_mapping=m["pso"]
                )

            return JsonResponse({"message": "Mapping saved successfully"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=405)

def get_saved_mapping(request):
    subject_id = request.GET.get("subject_id")

    mappings = COPSOMap.objects.filter(subject_id=subject_id)

    data = []

    for m in mappings:
        data.append({
            "co": m.co.id,
            "po": m.po_mapping,
            "pso": m.pso_mapping
        })

    return JsonResponse(data, safe=False)


def get_branches(request):
    branches = POPSO.objects.values_list('branch', flat=True).distinct()
    return JsonResponse(list(branches), safe=False)

def get_po_pso_sessions(request):
    try:
        sessions = POPSO.objects.values_list('session', flat=True).distinct()
        return JsonResponse(list(sessions), safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)



def get_po_pso(request):
    try:
        session = request.GET.get("session")
        branch = request.GET.get("branch")

        data = POPSO.objects.filter(
            session=session,
            branch=branch
        )

        po_list = []
        pso_list = []

        for item in data:
            if item.type == "PO":
                po_list.append({
                    "code": item.code,
                    "description": item.description
                })
            elif item.type == "PSO":
                pso_list.append({
                    "code": item.code,
                    "description": item.description
                })

        return JsonResponse({
            "po": po_list,
            "pso": pso_list
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


def download_po_pso_pdf(request):
    branch = request.GET.get("branch")
    session = request.GET.get("session")

    #  Fetch data
    pos = POPSO.objects.filter(branch=branch, session=session, type="PO")
    psos = POPSO.objects.filter(branch=branch, session=session, type="PSO")

    #  Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="PO_PSO_Details.pdf"'

    #  PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()

    content = []

    #  Title
    content.append(Paragraph("PO - PSO Details", styles['Title']))
    content.append(Spacer(1, 10))

    #  Branch & Session
    content.append(Paragraph(f"Branch: {branch}", styles['Normal']))
    content.append(Paragraph(f"Session: {session}", styles['Normal']))
    content.append(Spacer(1, 10))

    #  PO Section
    content.append(Paragraph("Program Outcomes (PO)", styles['Heading2']))
    content.append(Spacer(1, 5))

    for po in pos:
        content.append(Paragraph(f"{po.code}: {po.description}", styles['Normal']))
        content.append(Spacer(1, 5))

    content.append(Spacer(1, 10))

    #  PSO Section
    content.append(Paragraph("Program Specific Outcomes (PSO)", styles['Heading2']))
    content.append(Spacer(1, 5))

    for pso in psos:
        content.append(Paragraph(f"{pso.code}: {pso.description}", styles['Normal']))
        content.append(Spacer(1, 5))

    #  Build PDF
    doc.build(content)

    return response

def download_mapping_excel(request):
    branch = request.GET.get("branch")
    session = request.GET.get("session")
    subject_id = request.GET.get("subject_id")
    subject_name = request.GET.get("subject")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CO-PO-PSO Mapping"

    ws.append(["Branch", branch])
    ws.append(["Session", session])
    ws.append(["Subject", subject_name])
    ws.append([])

    po_list = list(POPSO.objects.filter(branch=branch, session=session, type="PO"))
    pso_list = list(POPSO.objects.filter(branch=branch, session=session, type="PSO"))

    mappings = COPSOMap.objects.filter(subject_id=subject_id)

    # Header
    header = ["CO"] + [po.code for po in po_list] + [pso.code for pso in pso_list]
    ws.append(header)

    # Rows
    for m in mappings:
        row = []

        row.append(str(m.co))  # or m.co.co_number

        for po in po_list:
            row.append(m.po_mapping.get(po.code, "-"))

        for pso in pso_list:
            row.append(m.pso_mapping.get(pso.code, "-"))

        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=co_po_pso_mapping.xlsx"

    wb.save(response)
    return response

